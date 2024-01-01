import os
import openpyxl

base_path = r"C:\torquecomparisonproducts"

# Get only the first-level subdirectories under base_path
subdirectories = [d for d in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, d))]

for subdirectory in subdirectories:
    subdirectory_path = os.path.join(base_path, subdirectory)

    # Get only the Excel files directly inside the subdirectory
    excel_files = [f for f in os.listdir(subdirectory_path) if (f.endswith('.xlsx') or f.endswith('.XLSX') or f.endswith('.xls'))]

    for filename in excel_files:
        combined_file_name = f"Combined_{filename}"

        # Convert filenames to lowercase for case-insensitive checks
        filename_lower = filename.lower()

        # Check if combined file already exists
        if combined_file_name.lower() in [f.lower() for f in excel_files]:
            print(f"'{combined_file_name}' already exists in '{subdirectory}'. Skipping processing for '{filename}'.")
            continue

        if not filename_lower.startswith('result_'):
            # Here we create results_filename based on the extension in filename to maintain case consistency
            base_name, ext = os.path.splitext(filename)
            results_filename = "Result_" + base_name + ext

            # Check if results file exists considering case-insensitive filenames
            if results_filename.lower() in [f.lower() for f in excel_files]:
                file_path = os.path.join(subdirectory_path, filename)
                results_file_path = os.path.join(subdirectory_path, results_filename)

                # Load the original file and its corresponding results file
                original_wb = openpyxl.load_workbook(file_path, read_only=True)  # Load in read-only mode for faster reading
                original_ws = original_wb.active
                results_wb = openpyxl.load_workbook(results_file_path, read_only=True)
                results_ws = results_wb.active

                # Create a new workbook for the combined data
                combined_wb = openpyxl.Workbook()
                modified_ws = combined_wb.active
                modified_ws.title = "Sheet1"

                # Copy data starting from the 50th record and only columns A and B
                for row in original_ws.iter_rows(min_row=50, max_col=2, values_only=True):  # Using values_only for better performance
                    modified_ws.append(row)

                # Copy the results data to a new sheet in the combined workbook
                new_ws = combined_wb.create_sheet(title="Sheet2")
                for row in results_ws.iter_rows(values_only=True):  # Using values_only for better performance
                    new_ws.append(row)

                # Search Sheet2 for the specific value
                value_to_subtract = None
                for row in new_ws.iter_rows(max_col=2):  # Limit to first two columns for optimization
                    cell_value = row[0].value
                    if cell_value and "Apply input [s]" in str(cell_value):
                        try:
                            value_to_subtract = float(row[1].value)  # The value is in the next cell
                            break  # Exit the loop once the value is found
                        except ValueError:
                            print(f"Error extracting value from the string: {cell_value}")

                # Subtract the found value from all values in column A of Sheet1
                if value_to_subtract:
                    for cell in modified_ws["A"]:
                        if cell.value:
                            cell.value -= value_to_subtract

                # Find the first non-zero value in column A of Sheet1, then find the next zero
                zero_index = None
                has_found_non_zero = False

                for idx, cell in enumerate(modified_ws["A"], 1):
                    if not has_found_non_zero:
                        if abs(cell.value or 0) > 1e-5:  # Check if non-zero
                            has_found_non_zero = True
                    else:
                        if abs(cell.value or 0) < 1e-5:  # Check if close to zero after a non-zero value
                            zero_index = idx
                            break

                # Delete all rows above the found zero index, but retain 5000 rows before it
                if zero_index and zero_index > 1:
                    starting_index_to_retain = zero_index - 5000 if zero_index > 5001 else 1  # ensure the starting index isn't negative or 0
                    modified_ws.delete_rows(1, starting_index_to_retain - 1)

                # Save the combined data to a new file in the same directory
                combined_file_path = os.path.join(subdirectory_path, combined_file_name)
                combined_wb.save(combined_file_path)

                print(f"Processed '{filename}' and '{results_filename}' into '{combined_file_path}'")
