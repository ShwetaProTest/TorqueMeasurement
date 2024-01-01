import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Specify the directory path
ready_folder = "Ready_Files"

def process_excel_file(file_path):
    # Check if the file is a valid Excel file
    if not is_valid_excel_file(file_path):
        print(f"'{file_path}' is not a valid Excel file.")
        return

    # Check if "ready_data" sheet is already present
    workbook = load_workbook(file_path)
    if "ready_data" in workbook.sheetnames:
        print(f"'ready_data' sheet already exists in {file_path}. Skipping.")
        return

    # Read data from "sheet1" (raw_data) and skip the first 49 rows
    sheet1_data = pd.read_excel(file_path, sheet_name="raw_data", skiprows=48)

    # Remove the 3rd column
    sheet1_data = sheet1_data.drop(columns=sheet1_data.columns[2])

    # Open the Excel file
    workbook = load_workbook(file_path)

    # Create the target sheet "ready_data"
    workbook.create_sheet("ready_data")
    sheet3 = workbook["ready_data"]

    # Append the data from "sheet1_data" to "ready_data"
    for row in dataframe_to_rows(sheet1_data, index=False, header=False):
        sheet3.append(row)

    # Get the "Apply Input Time" cell value from "sheet2" ("raw_analysed_data")
    sheet2 = workbook["raw_analysed_data"]
    apply_input_time = None
    for row in sheet2.iter_rows(values_only=True):
        for cell in row:
            if row[0] == "Apply input time [s]":
                apply_input_time = float(row[1])
                break

    if apply_input_time is not None:
        # Subtract the "Apply Input Time" value from the values in the first column of "ready_data"
        for row in sheet1_data.iterrows():
            for cell in row[1]:
                cell = sheet3.cell(row=row[0] + 2, column=1, value=cell - apply_input_time)

    # Find the first non-zero value in the 1st column
    first_non_zero_index = 0
    for row in sheet3.iter_rows(min_col=1, max_col=1, values_only=True):
        if row[0] != 0:
            break
        first_non_zero_index += 1

    # Delete all rows above the found zero index while retaining 5000 rows before it
    if first_non_zero_index > 5000:
        sheet3.delete_rows(1, first_non_zero_index - 5000)

    # Save the modified file
    workbook.save(file_path)

    print(f"Data from 'sheet1' has been appended to 'ready_data' and modified in the same file: {file_path}")

# Function to check if a file is a valid Excel file
def is_valid_excel_file(file_path):
    engines_to_try = ['openpyxl', 'xlrd', 'odf']
    for engine in engines_to_try:
        try:
            with pd.ExcelFile(file_path, engine=engine):
                return True
        except Exception as e:
            continue
    return False

# Traverse the directory and process Excel files
for root, _, files in os.walk(ready_folder):
    for file in files:
        file_path = os.path.join(root, file)
        process_excel_file(file_path)
